from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from selenium.webdriver.chrome.options import Options
import tkinter as tk
from tkinter import *
from tkinter.filedialog import askopenfilename
import os
import openpyxl
from tkinter.ttk import *


filename = ""
root=tk.Tk() 
root.title('TweetSnap')   
root.geometry('500x100')
# progress = Progressbar(root, orient = HORIZONTAL, length = 100, mode = 'determinate')
var = StringVar()
lb2 = Label(root, textvariable = var)


def browsefunc():
    global filename
    filename =tk.filedialog.askopenfilename(filetypes=(("xlsx files","*.xlsx"),("All files","*.*")))
    ent1.insert(tk.END, filename) # add this

def tweet_snap():
    if not os.path.exists('tweets'):
        os.mkdir('tweets')
    wb = openpyxl.load_workbook(filename)
    ws = wb['tweets']
    DRIVER_PATH = './chromedriver_linux64/chromedriver'
    options = Options()
    options.headless = True
    wd = webdriver.Chrome(executable_path=DRIVER_PATH, options= options)
    # progress.grid(row=2, column = 0)
    lb2.grid(row = 2, column = 0)
    for i in range(2, ws.max_row+1):
        # progress['value'] = int((i-1/(ws.max_row-1)) * 100)
        var.set('completed: '+str(i-1) + '/'+str(ws.max_row - 1))
        root.update_idletasks()
        link = ws.cell(row=i, column=1).hyperlink.target
        wd.get('https://tweetcyborg.com/')
        search_box = wd.find_element_by_class_name('url-input')
        search_box.send_keys(link)
        wd.find_element_by_class_name('download').click()
        time.sleep(1)
        element_present = EC.presence_of_element_located((By.CLASS_NAME, 'download'))
        WebDriverWait(wd, 20).until(element_present)
        wd.find_element_by_class_name('download').click()
        time.sleep(2)
        os.rename('tweetcyborg.png', 'tweets/'+link.split('/')[-1] + '.png')
        # break
    wd.quit()

def stop():
    root.destroy()
    


lb1 = Label(root, text = "Welcome, Please choose the excel file.", font="Helvetica 16 bold italic")
lb1.grid(row = 0, column = 0)
ent1=tk.Entry(root,font="Helvetica 16 bold italic")
ent1.grid(row=1,column=0)
b1=tk.Button(root,text="Browse",font="Helvetica 16 bold italic",command=browsefunc)
b1.grid(row=1,column=1)
b2 = tk.Button(root, text = "Start", font = "Helvetica 16 bold italic", command =tweet_snap)
b2.grid(row=3, column=0)
# b3 = tk.Button(root, text = "Exit", font = "Helvetica 16 bold italic", command =stop)
# b3.grid(row=3, column=1)
root.mainloop()

